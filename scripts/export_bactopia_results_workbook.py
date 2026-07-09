#!/usr/bin/env python3

from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
except ImportError as exc:  # pragma: no cover
    sys.stderr.write(
        "openpyxl is required for workbook export.\n"
        "Install it in the Python environment used on Gadi, for example:\n"
        "  python3 -m pip install --user openpyxl\n"
    )
    raise SystemExit(1) from exc


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--results-root")
    parser.add_argument("--consolidated-dir")
    parser.add_argument("--output", required=True)
    parser.add_argument("--st131typer-dir")
    parser.add_argument("--append", action="store_true")
    return parser.parse_args()


def read_table(path: Path) -> list[list[str]]:
    delimiter = "," if path.suffix.lower() == ".csv" else "\t"
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle, delimiter=delimiter)
        return [list(row) for row in reader]


def sanitize_sheet_name(name: str, used: set[str]) -> str:
    cleaned = name
    for char in "[]*?/\\:":
        cleaned = cleaned.replace(char, "_")
    while "__" in cleaned:
        cleaned = cleaned.replace("__", "_")
    cleaned = cleaned.strip("_") or "sheet"
    cleaned = cleaned[:31]
    candidate = cleaned
    counter = 1
    while candidate in used:
        suffix = f"_{counter}"
        candidate = f"{cleaned[: max(1, 31 - len(suffix))]}{suffix}"
        counter += 1
    used.add(candidate)
    return candidate


def collect_tables(dir_path: Path, prefix: str | None = None) -> list[tuple[str, Path]]:
    if not dir_path.exists():
        return []

    tables: list[tuple[str, Path]] = []
    for path in sorted(dir_path.rglob("*")):
        if path.suffix.lower() not in {".tsv", ".csv"} or not path.is_file():
            continue
        rel = path.relative_to(dir_path)
        sheet_name = str(rel.with_suffix("")).replace("/", "__")
        if prefix:
            sheet_name = f"{prefix}__{sheet_name}"
        tables.append((sheet_name, path))
    return tables


def choose_unique_results_file(results_root: Path, pattern: str) -> Path | None:
    matches = sorted(path for path in results_root.glob(pattern) if path.is_file())
    if not matches:
        return None
    if len(matches) > 1:
        joined = "\n".join(f"  {path}" for path in matches)
        raise SystemExit(
            f"Multiple results files matched {pattern} under {results_root}:\n{joined}\n"
            "Pass a single results set into the workbook export."
        )
    return matches[0]


def choose_mapped_results(results_root: Path) -> Path:
    preferred = choose_unique_results_file(results_root, "*_samplesheet_with_results_mlst_reviewed.tsv")
    fallback = choose_unique_results_file(results_root, "*_samplesheet_with_results.tsv")
    if preferred is not None:
        return preferred
    if fallback is not None:
        return fallback
    raise SystemExit(
        "Neither reviewed nor mapped metadata results were found. Expected one of:\n"
        f"  {results_root}/*_samplesheet_with_results_mlst_reviewed.tsv\n"
        f"  {results_root}/*_samplesheet_with_results.tsv\n"
    )


def mapped_sheet_name(mapped_path: Path) -> str:
    stem = mapped_path.stem
    for suffix in ("_with_results_mlst_reviewed", "_with_results"):
        if stem.endswith(suffix):
            return f"{stem[:-len(suffix)]}_mapped"
    return stem


def find_st131typer_summary(dir_path: Path) -> Path | None:
    candidates = [
        dir_path / "summary.txt",
        dir_path / "summary.tsv",
        dir_path / "summary.csv",
    ]
    for candidate in candidates:
        if candidate.exists() and candidate.is_file():
            return candidate

    for pattern in ("summary.txt*", "summary.tsv*", "summary.csv*"):
        matches = sorted(p for p in dir_path.glob(pattern) if p.is_file())
        if matches:
            return matches[0]

    return None


def add_sheet(workbook: Workbook, used_names: set[str], name: str, rows: list[list[str]]) -> None:
    worksheet = workbook.create_sheet(title=sanitize_sheet_name(name, used_names))
    for row in rows:
        worksheet.append(row)


def load_output_workbook(output_path: Path, append: bool) -> Workbook:
    if append and output_path.exists():
        return load_workbook(output_path)

    workbook = Workbook()
    workbook.remove(workbook.active)
    return workbook


def main() -> int:
    args = parse_args()
    results_root = Path(args.results_root) if args.results_root else None
    consolidated_dir = Path(args.consolidated_dir) if args.consolidated_dir else None
    output_path = Path(args.output)
    st131typer_dir = Path(args.st131typer_dir) if args.st131typer_dir else None

    if consolidated_dir is not None and results_root is None:
        raise SystemExit("--results-root is required when --consolidated-dir is provided")

    if not args.append and (results_root is None or consolidated_dir is None):
        raise SystemExit("--results-root and --consolidated-dir are required unless --append is used")

    if results_root is not None and not results_root.is_dir():
        raise SystemExit(f"--results-root must point to an existing directory: {results_root}")
    if consolidated_dir is not None and not consolidated_dir.is_dir():
        raise SystemExit(f"--consolidated-dir must point to an existing directory: {consolidated_dir}")
    if st131typer_dir is not None and not st131typer_dir.is_dir():
        raise SystemExit(f"--st131typer-dir must point to an existing directory: {st131typer_dir}")

    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = load_output_workbook(output_path, args.append)
    used_names: set[str] = set(workbook.sheetnames)

    if results_root is not None and consolidated_dir is not None:
        mapped_path = choose_mapped_results(results_root)
        add_sheet(workbook, used_names, mapped_sheet_name(mapped_path), read_table(mapped_path))

        for top_level in ("project_summary.tsv", "tool_processing_log.tsv"):
            path = consolidated_dir / top_level
            if path.exists():
                add_sheet(workbook, used_names, path.stem, read_table(path))

        for sheet_name, path in collect_tables(consolidated_dir / "results_main" / "merged-results", "main"):
            add_sheet(workbook, used_names, sheet_name, read_table(path))

        for sheet_name, path in collect_tables(consolidated_dir / "tools", "tool"):
            add_sheet(workbook, used_names, sheet_name, read_table(path))

    if st131typer_dir is not None:
        summary_path = find_st131typer_summary(st131typer_dir)
        if summary_path is None:
            raise SystemExit(
                f"No ST131Typer summary table was found under: {st131typer_dir}. "
                "Expected summary.txt, summary.tsv, or summary.csv."
            )
        add_sheet(workbook, used_names, "st131typer_summary", read_table(summary_path))

    workbook.save(output_path)
    print(f"Excel workbook written with openpyxl: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
